from openpyxl import Workbook


# pensar usar clase VistaGeneral
class Salida:
    def __init__(self, archivos_con_diferencias, directorio_salida):
        self.archivos_con_diferencias = archivos_con_diferencias
        self.directorio_salida = directorio_salida
        self.wb = Workbook()

    def _generar_header(self, ws, hoja):
        ws.append(['Archivo', f'Diferencias encontradas en {hoja}', ])

    def _generar_hojas(self):
        for archivo in self.archivos_con_diferencias:
            for diferencias in self.archivos_con_diferencias[archivo]:
                hoja = diferencias.obtener_tipo_pagina()
                if hoja not in self.wb.sheetnames:
                    ws = self.wb.create_sheet(title=hoja)

    def generar_salida(self):
        self._generar_hojas()
        self.wb.save(self.directorio_salida)
